# =============================================================================
# CONSUMER SCORECARD V4 — PYSPARK VALIDATION SCRIPT
# =============================================================================
#
# PURPOSE
# -------
# Compares pipeline output CSV(s) against a ground-truth reference file and
# produces a colour-coded Excel validation report identical in logic to the
# hand-built Variable_Validation_Report but fully automated.
#
# USAGE
# -----
#   python scorecard_validator.py \
#       --reference  /path/to/reference.csv \
#       --pipeline   /path/to/pipeline_output.csv \
#       --output     /path/to/validation_report.xlsx \
#       [--label     "My run label"] \
#       [--exact-tol  0.0001] \
#       [--near-tol   5.0] \
#       [--skip-cols  col1 col2 col3]
#
# ALWAYS-SKIPPED COLUMNS (pipeline scoring cols, never in ground truth)
# ---------------------------------------------------------------------
#   scorecard_name       — scorecard tier (ST_1_SE, ST_2, etc.)
#   final_score          — numeric credit score 300–900
#   score_breakdown      — pipe-delimited score contribution string
#   is_eligible_for_st2  — boolean trigger flag
#   is_eligible_for_st3  — boolean trigger flag
#
# These 5 are skipped automatically every run. You do NOT need to pass them
# via --skip-cols. To skip additional columns use:
#
#   --skip-cols  util_l3m_uns_tot max_simul_pl_cd
#
# INPUTS
# ------
#   reference  : CSV or Excel with one row per customer.
#                Columns = cust_id + any scorecard variable names.
#                Can also point at a multi-row Excel sheet with reference blocks
#                (same format as Variable.xlsx — detected automatically).
#
#   pipeline   : CSV produced by run_pipeline() in con10_fixed.py.
#                Also accepts Excel.
#
# OUTPUT
# ------
#   Excel workbook with 4 sheets:
#     1. Dashboard      — summary KPIs + accuracy bar per customer
#     2. Detail         — one row per variable, per customer; colour-coded
#     3. Mismatches     — only failing variables with root-cause hints
#     4. Column Audit   — checks for duplicate / missing / extra columns
#
# THRESHOLDS
# ----------
#   exact  : |ref - code| < exact_tol (default 0.0001)
#   near   : relative % diff < near_tol (default 5%)
#   NaN    : both null  → ✔  |  ref null code has value → ⚠  |  inverse → ⚠
#
# EXAMPLES
# --------
#   # Basic — uses ground_truth.csv and scorecard output:
#   python scorecard_validator.py \
#       --reference ground_truth.csv \
#       --pipeline  scorecard_output_20250930_143022.csv \
#       --output    validation_report.xlsx
#
#   # With a custom label:
#   python scorecard_validator.py \
#       --reference Variable.xlsx \
#       --pipeline  scorecard_output_20250930_143022.csv \
#       --output    report.xlsx \
#       --label     "Trade.xlsx run — new code"
#
#   # Skip additional columns beyond the 5 built-in ones:
#   python scorecard_validator.py \
#       --reference ground_truth.csv \
#       --pipeline  output.csv \
#       --skip-cols util_l3m_uns_tot max_simul_pl_cd
#
#   # Strict mode (1% near tolerance):
#   python scorecard_validator.py \
#       --reference ground_truth.csv \
#       --pipeline  output.csv \
#       --near-tol  1.0
#
#   # PySpark mode for large files (>100k customers):
#   python scorecard_validator.py \
#       --reference ground_truth.csv \
#       --pipeline  output.csv \
#       --spark
#
# =============================================================================

import os
import sys
import math
import argparse
import datetime
from collections import Counter, defaultdict

import pandas as pd
import numpy as np

# ── Optional PySpark (used for large reference files; falls back to pandas) ──
try:
    from pyspark.sql import SparkSession
    from pyspark.sql import functions as F
    from pyspark.sql.types import StringType, DoubleType, IntegerType, StructType, StructField
    SPARK_AVAILABLE = True
except ImportError:
    SPARK_AVAILABLE = False

# ── openpyxl for Excel output ─────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARN] openpyxl not found — install with: pip install openpyxl")


# =============================================================================
# SECTION 1 — CONFIGURATION
# =============================================================================

EXACT_TOL_DEFAULT  = 1e-4   # |ref - code| < this → Exact Match
NEAR_PCT_DEFAULT   = 5.0    # relative % diff < this → Near Match

# Pipeline-only columns that are ALWAYS skipped — they exist in the pipeline
# output but never appear in any ground-truth reference file.
# Add more names here if the pipeline grows new scoring/meta columns.
PIPELINE_ONLY_COLS = {
    "scorecard_name",       # scorecard tier assigned (ST_1_SE, ST_2, etc.)
    "final_score",          # numeric credit score 300–900
    "score_breakdown",      # pipe-delimited string of score contributions
    "is_eligible_for_st2",  # boolean eligibility flag
    "is_eligible_for_st3",  # boolean eligibility flag
}

# Variables where 0 is meaningful (not NaN) — do not flag 0 vs NaN as error
ZERO_IS_VALID = {
    "nbr_0_0m_all", "nbr_0_24m_live", "nbr_cc40l6m_tot_accts_36",
    "totconsinc_bal10_exc_cc_25m", "totconsinc_bal10_exc_cc_7m",
    "totconsinc_bal5_exc_cc_7m", "totconsinc_util1_tot_7m",
    "consinc_bal10_exc_cc_25m", "consinc_bal10_tot_7m",
    "consinc_bal5_cc_13m", "totconsdec_bal_tot_36m",
    "freq_between_accts_unsec_wo_cc",
    "nbr_accts_open_l6m", "nbr_accts_open_7to12m", "nbr_accts_open_l12m_wo_cc",
    "HL_LAP_outflow", "PL_outflow", "AL_TW_outflow", "CD_outflow",
}

# Known formula/direction issues — flag as KNOWN_ISSUE not pure mismatch
KNOWN_ISSUES = {
    "freq_between_accts_all":            "Formula direction: avg gap vs rate — under investigation",
    "freq_between_accts_unsec_wo_cc":    "Formula direction: avg gap vs rate — under investigation",
    "freq_between_installment_trades":   "Formula direction: avg gap vs rate — under investigation",
    "balance_amt_0_12_by_13_24":         "Sec account inclusion: HL drives 13-24m window",
    "util_l12m_uns_tot":                 "Near-match (~4% gap) — per-account formula precision",
    "consinc_bal5_cc_13m":               "Loop start: accounts reporting from month_diff=2+ miss idx=0",
    "mon_since_max_bal_l24m_uns":        "Off by 1 — ordering/liveness edge case",
}

# Variable groups for organising the Detail sheet
VARIABLE_GROUPS = {
    "Account Counts": [
        "nbr_tot_accts_36","nbr_live_accts_36","nbr_al_tot_accts_36","nbr_hl_tot_accts_36",
        "nbr_cc_tot_accts_36","nbr_cc_live_accts_36","nbr_agri_tot_accts_36",
        "nbr_comsec_tot_accts_36","nbr_comuns_tot_accts_36","nbr_pl_tot_accts_36",
        "nbr_tw_tot_accts_36","nbr_cl_tot_accts_36","nbr_pl_le50_tot_accts_36",
        "nbr_cc40l6m_tot_accts_36",
    ],
    "MOB & Recency": [
        "max_mob_all_36","min_mob_all_36","avg_mob_reg_36","max_mob_cc_36",
        "min_mob_uns_wo_cc_36","min_mob_agri_live_36","max_mob_comuns_live_36",
        "mon_since_last_acct_open","mon_since_max_bal_l24m_uns",
        "nbr_accts_open_l6m","nbr_accts_open_7to12m","nbr_accts_open_l12m_wo_cc",
        "open_cnt_0_6_by_7_12",
    ],
    "Simultaneous Accounts": [
        "max_simul_unsec_wo_cc","max_simul_unsec_wo_cc_inc_gl",
        "max_simul_unsec","max_simul_pl_cd",
    ],
    "DPD & Delinquency": [
        "max_dpd__UNS_L36M","max_dpd_L30M","max_dpd__UNS__L6M",
        "max_dpd__UNS_L_6_12_M","max_dpd__SEC__0_live",
        "nbr_0_0m_all","nbr_0_24m_live","max_nbr_0_24m_uns",
        "mon_sin_recent_1","mon_sin_recent_2","mon_sin_recent_3","mon_sin_recent_4",
        "mon_sin_recent_5","mon_sin_recent_6","mon_sin_recent_7",
        "mon_sin_first_1","mon_sin_first_2","mon_sin_first_3","mon_sin_first_4",
        "mon_sin_first_5","mon_sin_first_6","mon_sin_first_7",
    ],
    "Utilisation": [
        "util_l3m_cc_live","util_l6m_all_tot","util_l12m_all_tot",
        "avg_util_l6m_all_tot","util_l3m_uns_tot","util_l6m_uns_tot",
        "util_l12m_uns_tot","util_l3m_exc_cc_live",
    ],
    "Sanction Amounts": [
        "max_sanc_amt","max_sanc_amt_sec","max_sanc_amt_secmov",
        "max_sanc_amt_pl","max_sanc_amt_al","max_sanc_amt_tw",
        "max_sanc_amt_cl","sum_sanc_amt_uns",
    ],
    "Outflows": [
        "total_monthly_outflow_wo_cc","HL_LAP_outflow",
        "AL_TW_outflow","PL_outflow","CD_outflow",
    ],
    "Balance Trends": [
        "totconsinc_bal10_exc_cc_25m","totconsdec_bal_tot_36m",
        "totconsinc_bal10_exc_cc_7m","totconsinc_bal5_exc_cc_7m",
        "consinc_bal10_exc_cc_25m","consinc_bal10_tot_7m",
        "totconsinc_util1_tot_7m","consinc_bal5_cc_13m",
    ],
    "Balance Ratios": [
        "balance_amt_0_12_by_13_24","balance_amt_0_6_by_7_12",
        "avg_b_1to6_by_7_12_agri","delinq_bal_0_12_by_13_24",
        "delinq_bal_13_24_by_25_36","live_cnt_0_6_by_7_12",
    ],
    "Frequency": [
        "freq_between_accts_all","freq_between_accts_unsec_wo_cc",
        "freq_between_installment_trades",
    ],
    "Other": ["latest_account_type"],
}


# =============================================================================
# SECTION 2 — DATA LOADING
# =============================================================================

def _read_file(path: str) -> pd.DataFrame:
    """Load CSV or Excel into pandas. Handles byte-string wrapped values."""
    ext = os.path.splitext(path)[-1].lower()
    if ext in ('.xlsx', '.xls', '.xlsm'):
        df = pd.read_excel(path, header=0)
    else:
        df = pd.read_csv(path, low_memory=False)
    # Strip b'' byte-string wrapping that some Spark exports add
    for c in df.select_dtypes(include='object').columns:
        df[c] = df[c].astype(str).str.replace(r"^b'|'$", "", regex=True).str.strip()
        df[c] = df[c].replace({'nan': np.nan, 'None': np.nan, '': np.nan})
    return df


def load_reference(path: str) -> dict[str, pd.Series]:
    """
    Load reference file. Returns {cust_id_str: pd.Series of variable→value}.
    Handles both:
      - Simple CSV: one header row + one data row per customer
      - Variable.xlsx format: multi-block with label rows (auto-detected)
    """
    ext = os.path.splitext(path)[-1].lower()

    if ext in ('.xlsx', '.xls', '.xlsm'):
        raw = pd.read_excel(path, sheet_name=0, header=None)
        # Detect multi-block format: label cell is in col 0 or col 1
        # Variable.xlsx uses col 1 (col 0 is empty/index)
        def _has_ref(row_idx):
            for col in range(min(3, raw.shape[1])):
                if 'Reference' in str(raw.iloc[row_idx, col]):
                    return True
            return False
        ref_rows = [i for i in range(raw.shape[0]) if _has_ref(i)]
        if ref_rows:
            return _parse_variable_xlsx(raw)

    # Simple format: first row = headers, remaining = data
    df = _read_file(path)
    if 'cust_id' not in df.columns:
        raise ValueError(f"Reference file must have a 'cust_id' column. Found: {list(df.columns)}")
    result = {}
    for _, row in df.iterrows():
        cid = str(int(float(row['cust_id'])))
        result[cid] = row.drop('cust_id')
    return result


def _parse_variable_xlsx(raw: pd.DataFrame) -> dict[str, pd.Series]:
    """Parse the multi-block Variable.xlsx format used in this project."""
    result = {}
    i = 0
    # Find which column holds the block labels (usually col 0 or col 1)
    label_col = 0
    for _c in range(min(3, raw.shape[1])):
        if any('Reference' in str(raw.iloc[_r, _c]) for _r in range(raw.shape[0])):
            label_col = _c; break

    while i < raw.shape[0]:
        cell0 = str(raw.iloc[i, label_col]).strip()
        if cell0.startswith('Reference'):
            # Next row = header, row after = data
            if i + 2 < raw.shape[0]:
                headers = raw.iloc[i+1].tolist()
                values  = raw.iloc[i+2].tolist()
                d = {}
                seen = set()
                for h, v in zip(headers, values):
                    h = str(h).strip()
                    if h in ('nan', 'None', '') or h in seen:
                        continue
                    seen.add(h)
                    if str(v).strip() in ('nan', 'None', ''):
                        v = None
                    else:
                        try: v = float(v)
                        except (TypeError, ValueError): pass
                    d[h] = v
                cid = str(int(float(d.get('cust_id', 0) or 0)))
                result[cid] = pd.Series(d).drop('cust_id', errors='ignore')
                i += 3
                continue
        i += 1
    return result


def load_pipeline_output(path: str, skip_cols: set | None = None) -> dict[str, pd.Series]:
    """Load pipeline output CSV/Excel. Returns {cust_id_str: pd.Series}.
    Columns in PIPELINE_ONLY_COLS and any extra skip_cols are dropped before comparison.
    """
    df = _read_file(path)
    if 'cust_id' not in df.columns:
        raise ValueError(f"Pipeline output must have a 'cust_id' column. Found: {list(df.columns)}")

    # Drop pipeline-only scoring columns (always) + any caller-supplied extras
    cols_to_drop = PIPELINE_ONLY_COLS | (skip_cols or set())
    cols_to_drop = {c for c in cols_to_drop if c in df.columns}
    if cols_to_drop:
        print(f"  Skipping {len(cols_to_drop)} pipeline-only/excluded columns: {sorted(cols_to_drop)}")
        df = df.drop(columns=list(cols_to_drop))

    # Deduplicate columns — keep first occurrence
    seen = set(); keep = []
    for c in df.columns:
        if c not in seen: keep.append(c); seen.add(c)
    if len(keep) < len(df.columns):
        dupes = [c for c in df.columns if list(df.columns).count(c) > 1]
        print(f"  [WARN] Duplicate columns in pipeline output (kept first): {set(dupes)}")
        df = df[keep]

    # Convert all numeric columns
    for c in df.columns:
        if c == 'cust_id': continue
        try:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        except Exception:
            pass

    result = {}
    for _, row in df.iterrows():
        try:
            cid = str(int(float(row['cust_id'])))
        except (ValueError, TypeError):
            cid = str(row['cust_id'])
        result[cid] = row.drop('cust_id')
    return result


# =============================================================================
# SECTION 3 — COMPARISON ENGINE
# =============================================================================

def is_null(v) -> bool:
    if v is None: return True
    try:
        f = float(v)
        return math.isnan(f)
    except (TypeError, ValueError):
        return str(v).strip().lower() in ('nan', 'none', '', 'null')


def classify_pair(ref_val, code_val, col: str, exact_tol: float, near_pct: float) -> dict:
    """
    Classify a (ref, code) pair into one of:
      exact | near | mismatch | ref_nan | code_nan | both_nan | known_issue
    Returns a dict with keys: status, ref, code, pct_diff, note
    """
    rn, cn = is_null(ref_val), is_null(code_val)

    note = KNOWN_ISSUES.get(col, "")

    if rn and cn:
        return dict(status='both_nan', ref=None, code=None, pct_diff=None, note=note)
    if rn and not cn:
        return dict(status='ref_nan',  ref=None, code=code_val, pct_diff=None, note=note)
    if cn and not rn:
        return dict(status='code_nan', ref=ref_val, code=None, pct_diff=None, note=note)

    # Both have values — try numeric comparison
    try:
        r, c = float(ref_val), float(code_val)
        abs_diff = abs(r - c)
        pct_diff = (abs_diff / abs(r) * 100) if r != 0 else (abs(c) * 100)

        if abs_diff <= exact_tol:
            status = 'exact'
        elif pct_diff <= near_pct:
            status = 'near'
        elif note:
            status = 'known_issue'
        else:
            status = 'mismatch'

        return dict(status=status, ref=r, code=c, pct_diff=pct_diff, note=note)

    except (TypeError, ValueError):
        # String comparison
        rs, cs = str(ref_val).strip(), str(code_val).strip()
        status = 'exact' if rs == cs else 'mismatch'
        return dict(status=status, ref=rs, code=cs, pct_diff=None, note=note)


STATUS_RANK = {
    'exact': 0, 'near': 1, 'both_nan': 2,
    'known_issue': 3, 'ref_nan': 4, 'code_nan': 5, 'mismatch': 6,
}
IS_MATCH = {'exact', 'near', 'both_nan'}
IS_WARNING = {'ref_nan', 'code_nan', 'known_issue'}
IS_FAIL = {'mismatch'}


def compare_customer(ref_series: pd.Series, code_series: pd.Series,
                     exact_tol: float, near_pct: float) -> list[dict]:
    """Compare all variables for one customer. Returns list of result dicts."""
    all_cols = sorted(set(ref_series.index) | set(code_series.index))
    results = []
    for col in all_cols:
        rv = ref_series.get(col)
        cv = code_series.get(col)
        r = classify_pair(rv, cv, col, exact_tol, near_pct)
        r['col'] = col
        # Determine group
        r['group'] = next(
            (g for g, cols in VARIABLE_GROUPS.items() if col in cols),
            'Other'
        )
        results.append(r)
    return results


def summarise(results: list[dict]) -> dict:
    total   = len(results)
    match   = sum(1 for r in results if r['status'] in IS_MATCH)
    warn    = sum(1 for r in results if r['status'] in IS_WARNING)
    fail    = sum(1 for r in results if r['status'] in IS_FAIL)
    known   = sum(1 for r in results if r['status'] == 'known_issue')
    exact   = sum(1 for r in results if r['status'] == 'exact')
    near    = sum(1 for r in results if r['status'] == 'near')
    ref_nan = sum(1 for r in results if r['status'] == 'ref_nan')
    code_nan= sum(1 for r in results if r['status'] == 'code_nan')
    mismatch= sum(1 for r in results if r['status'] == 'mismatch')
    pct     = match / total * 100 if total else 0
    return dict(total=total, match=match, warn=warn, fail=fail, known=known,
                exact=exact, near=near, ref_nan=ref_nan, code_nan=code_nan,
                mismatch=mismatch, pct=pct)


# =============================================================================
# SECTION 4 — COLUMN AUDIT
# =============================================================================

def audit_columns(ref_cols: list[str], code_cols: list[str]) -> dict:
    ref_set  = set(ref_cols)
    code_set = set(code_cols)
    missing  = sorted(ref_set  - code_set)   # in ref but not code
    extra    = sorted(code_set - ref_set)     # in code but not ref
    common   = sorted(ref_set  & code_set)
    dupes_code = [c for c, n in Counter(code_cols).items() if n > 1]
    return dict(missing=missing, extra=extra, common=common,
                dupes_code=sorted(dupes_code))


# =============================================================================
# SECTION 5 — EXCEL REPORT BUILDER
# =============================================================================

# Colour palette
PAL = {
    'exact_bg':  'D6F0DC', 'exact_fg':  '155724',
    'near_bg':   'E8F8E8', 'near_fg':   '276221',
    'nan_bg':    'F5F5F5', 'nan_fg':    '666666',
    'warn_bg':   'FFF3CD', 'warn_fg':   '856404',
    'known_bg':  'E8E0FF', 'known_fg':  '4B0082',
    'fail_bg':   'FFE4E4', 'fail_fg':   '8B0000',
    'head_bg':   '0D3C6E', 'head_fg':   'FFFFFF',
    'sub_bg':    '1A5276', 'sub_fg':    'FFFFFF',
    'group_bg':  'D6E8F7', 'group_fg':  '0D3C6E',
    'white':     'FFFFFF', 'light':     'F8F9FA',
    'dark':      '1F1F1F',
}

STATUS_STYLE = {
    'exact':       ('exact_bg',  'exact_fg',  '✔ Exact'),
    'near':        ('near_bg',   'near_fg',   '≈ Near'),
    'both_nan':    ('nan_bg',    'nan_fg',    '✔ Both NaN'),
    'known_issue': ('known_bg',  'known_fg',  '⚠ Known'),
    'ref_nan':     ('warn_bg',   'warn_fg',   '⚠ Ref=NaN'),
    'code_nan':    ('warn_bg',   'warn_fg',   '⚠ Code=NaN'),
    'mismatch':    ('fail_bg',   'fail_fg',   '✘ Mismatch'),
}


def _thin():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bg='FFFFFF', fg='1F1F1F',
          bold=False, align='center', sz=10, wrap=False, fmt=None):
    c = ws.cell(row, col)
    c.value = value
    c.font  = Font(bold=bold, color=fg, name='Calibri', size=sz)
    c.fill  = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = _thin()
    if fmt:
        c.number_format = fmt
    elif isinstance(value, float) and not math.isnan(value):
        c.number_format = '#,##0.000000'
    return c


def _header(ws, row, col, text, span=1, bg=None, fg=None, sz=11):
    bg = bg or PAL['head_bg']; fg = fg or PAL['head_fg']
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row, col, text)
    c.font = Font(bold=True, color=fg, name='Calibri', size=sz)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _thin()
    ws.row_dimensions[row].height = 22
    return c


def _col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _status_cell(ws, row, col, status):
    bg_key, fg_key, label = STATUS_STYLE.get(status, ('white', 'dark', status))
    bg = PAL[bg_key]; fg = PAL[fg_key]
    return _cell(ws, row, col, label, bg=bg, fg=fg, bold=True, sz=9)


def fmt_val(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return '—'
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e9:
            return str(int(v))
        return f"{v:.6f}"
    return str(v)


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────

def build_dashboard(wb, all_results: dict[str, list], label: str, run_ts: str):
    ws = wb.active
    ws.title = '📊 Dashboard'
    ws.sheet_view.showGridLines = False
    _col_w(ws, [30, 14, 14, 14, 14, 14, 14, 14])

    # Title
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'Consumer Scorecard V4 — Validation Report  |  {label}  |  {run_ts}'
    c.font  = Font(bold=True, name='Calibri', size=16, color=PAL['head_fg'])
    c.fill  = PatternFill('solid', start_color=PAL['head_bg'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    # Column headers
    row = 2
    hdrs = ['Customer', 'Total Vars', '✔ Exact', '≈ Near', '✔ Match%',
            '✘ Mismatch', '⚠ Warning', '⚠ Known Issue']
    for ci, h in enumerate(hdrs, 1):
        _header(ws, row, ci, h, sz=10)

    # One row per customer
    cust_ids = sorted(all_results.keys())
    for ri, cid in enumerate(cust_ids, 3):
        results = all_results[cid]
        s = summarise(results)
        bg = PAL['light'] if ri % 2 == 0 else PAL['white']
        pct_bg = PAL['exact_bg'] if s['pct'] >= 80 else (PAL['warn_bg'] if s['pct'] >= 60 else PAL['fail_bg'])
        pct_fg = PAL['exact_fg'] if s['pct'] >= 80 else (PAL['warn_fg'] if s['pct'] >= 60 else PAL['fail_fg'])

        ws.row_dimensions[ri].height = 20
        _cell(ws, ri, 1, f"cust {cid}", bg=bg, align='left', bold=True, sz=11)
        _cell(ws, ri, 2, s['total'],  bg=bg, sz=11)
        _cell(ws, ri, 3, s['exact'],  bg=PAL['exact_bg'], fg=PAL['exact_fg'], bold=True, sz=11)
        _cell(ws, ri, 4, s['near'],   bg=PAL['near_bg'],  fg=PAL['near_fg'],  bold=True, sz=11)
        _cell(ws, ri, 5, f"{s['pct']:.1f}%  ({s['match']}/{s['total']})",
              bg=pct_bg, fg=pct_fg, bold=True, sz=12)
        _cell(ws, ri, 6, s['mismatch'],  bg=PAL['fail_bg'],  fg=PAL['fail_fg'],  bold=True, sz=11)
        _cell(ws, ri, 7, s['ref_nan'] + s['code_nan'],
              bg=PAL['warn_bg'],  fg=PAL['warn_fg'],  bold=True, sz=11)
        _cell(ws, ri, 8, s['known'],  bg=PAL['known_bg'], fg=PAL['known_fg'], bold=True, sz=11)

    # Totals row
    total_row = 3 + len(cust_ids)
    ws.row_dimensions[total_row].height = 22
    all_flat = [r for rs in all_results.values() for r in rs]
    s_all = summarise(all_flat)
    _header(ws, total_row, 1, '▶ OVERALL', bg=PAL['sub_bg'])
    _cell(ws, total_row, 2, s_all['total'],   bg=PAL['head_bg'], fg=PAL['head_fg'], bold=True, sz=11)
    _cell(ws, total_row, 3, s_all['exact'],   bg=PAL['exact_bg'], fg=PAL['exact_fg'], bold=True, sz=11)
    _cell(ws, total_row, 4, s_all['near'],    bg=PAL['near_bg'],  fg=PAL['near_fg'],  bold=True, sz=11)
    pct = s_all['pct']
    pct_bg = PAL['exact_bg'] if pct >= 80 else (PAL['warn_bg'] if pct >= 60 else PAL['fail_bg'])
    pct_fg = PAL['exact_fg'] if pct >= 80 else (PAL['warn_fg'] if pct >= 60 else PAL['fail_fg'])
    _cell(ws, total_row, 5, f"{pct:.1f}%  ({s_all['match']}/{s_all['total']})",
          bg=pct_bg, fg=pct_fg, bold=True, sz=12)
    _cell(ws, total_row, 6, s_all['mismatch'],  bg=PAL['fail_bg'],  fg=PAL['fail_fg'],  bold=True, sz=11)
    _cell(ws, total_row, 7, s_all['ref_nan'] + s_all['code_nan'],
          bg=PAL['warn_bg'],  fg=PAL['warn_fg'],  bold=True, sz=11)
    _cell(ws, total_row, 8, s_all['known'],  bg=PAL['known_bg'], fg=PAL['known_fg'], bold=True, sz=11)

    # Progress bars
    bar_row = total_row + 2
    for cid in cust_ids:
        s = summarise(all_results[cid])
        pct = s['pct'] / 100
        filled = int(pct * 40)
        bar = '█' * filled + '░' * (40 - filled)
        ws.merge_cells(start_row=bar_row, start_column=1, end_row=bar_row, end_column=8)
        c = ws.cell(bar_row, 1)
        c.value = f"cust {cid}  [{bar}]  {s['pct']:.1f}%  ({s['match']}/{s['total']} variables match)"
        c.font  = Font(name='Courier New', size=9, color=PAL['exact_fg'])
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[bar_row].height = 15
        bar_row += 1

    # Legend
    bar_row += 1
    legend_items = [
        ('✔ Exact',    PAL['exact_bg'],  PAL['exact_fg'],  '|ref − code| < 0.0001'),
        ('≈ Near',     PAL['near_bg'],   PAL['near_fg'],   'relative diff < 5%'),
        ('✔ Both NaN', PAL['nan_bg'],    PAL['nan_fg'],    'both ref and code are null'),
        ('⚠ Ref=NaN',  PAL['warn_bg'],   PAL['warn_fg'],   'ref is null but code has value'),
        ('⚠ Code=NaN', PAL['warn_bg'],   PAL['warn_fg'],   'code is null but ref has value'),
        ('⚠ Known',    PAL['known_bg'],  PAL['known_fg'],  'known formula/direction issue'),
        ('✘ Mismatch', PAL['fail_bg'],   PAL['fail_fg'],   'values differ beyond tolerance'),
    ]
    _header(ws, bar_row, 1, 'LEGEND', span=2, bg=PAL['sub_bg'])
    _header(ws, bar_row, 3, 'Description', span=6, bg=PAL['sub_bg'])

    for li, (lbl, bg, fg, desc) in enumerate(legend_items, bar_row + 1):
        ws.row_dimensions[li].height = 18
        _cell(ws, li, 1, lbl, bg=bg, fg=fg, bold=True)
        ws.merge_cells(start_row=li, start_column=3, end_row=li, end_column=8)
        _cell(ws, li, 3, desc, bg=PAL['white'], fg=PAL['dark'], align='left')


# ── Sheet 2: Detail ───────────────────────────────────────────────────────────

def build_detail(wb, all_results: dict[str, list], cust_ids: list[str]):
    ws = wb.create_sheet('📋 Detail')
    ws.sheet_view.showGridLines = False

    n_custs = len(cust_ids)
    # Columns: Variable | Group | (ref | code | status | pct_diff) × n_custs
    base_cols   = 2
    per_cust    = 4  # ref, code, status, pct_diff
    total_cols  = base_cols + n_custs * per_cust

    widths = [38, 18]
    for _ in cust_ids:
        widths += [16, 16, 14, 10]
    _col_w(ws, widths)

    # Row 1: merged title
    _header(ws, 1, 1, 'Variable Detail — All Customers', span=total_cols, sz=13)

    # Row 2: customer group headers
    ws.row_dimensions[2].height = 20
    _cell(ws, 2, 1, 'Variable', bg=PAL['head_bg'], fg=PAL['head_fg'], bold=True, align='left')
    _cell(ws, 2, 2, 'Group',    bg=PAL['head_bg'], fg=PAL['head_fg'], bold=True)
    for ci, cid in enumerate(cust_ids):
        base = base_cols + ci * per_cust + 1
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base+per_cust-1)
        _header(ws, 2, base, f'cust {cid}', span=per_cust, bg=PAL['sub_bg'], sz=10)

    # Row 3: sub-headers
    ws.row_dimensions[3].height = 18
    _cell(ws, 3, 1, '', bg=PAL['head_bg'])
    _cell(ws, 3, 2, '', bg=PAL['head_bg'])
    for ci in range(n_custs):
        base = base_cols + ci * per_cust + 1
        for j, lbl in enumerate(['Reference', 'Code Output', 'Status', '% Diff']):
            _cell(ws, 3, base+j, lbl, bg=PAL['head_bg'], fg=PAL['head_fg'], bold=True, sz=9)

    # Collect all variables in group order
    grouped_vars = []
    for grp, vars_in_grp in VARIABLE_GROUPS.items():
        for v in vars_in_grp:
            grouped_vars.append((grp, v))
    # Add any variables not in any group
    all_vars_in_groups = {v for _, v in grouped_vars}
    all_vars = set()
    for results in all_results.values():
        for r in results: all_vars.add(r['col'])
    for v in sorted(all_vars - all_vars_in_groups):
        grouped_vars.append(('Other', v))

    cur_group = None
    data_row = 4
    for grp, var in grouped_vars:
        if var not in all_vars:
            continue

        # Group header row
        if grp != cur_group:
            cur_group = grp
            ws.merge_cells(start_row=data_row, start_column=1,
                           end_row=data_row, end_column=total_cols)
            c = ws.cell(data_row, 1, f'  {grp}')
            c.font  = Font(bold=True, color=PAL['group_fg'], name='Calibri', size=10)
            c.fill  = PatternFill('solid', start_color=PAL['group_bg'])
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = _thin()
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # Determine worst status across all customers for row background
        statuses = []
        for cid in cust_ids:
            res_map = {r['col']: r for r in all_results.get(cid, [])}
            if var in res_map:
                statuses.append(res_map[var]['status'])

        worst = max(statuses, key=lambda s: STATUS_RANK.get(s, 0)) if statuses else 'both_nan'
        row_bg = PAL['white']
        if worst in IS_FAIL:    row_bg = PAL['fail_bg']
        elif worst == 'known_issue': row_bg = PAL['known_bg']
        elif worst in IS_WARNING: row_bg = PAL['warn_bg']
        elif worst in IS_MATCH:  row_bg = PAL['white']

        ws.row_dimensions[data_row].height = 16
        _cell(ws, data_row, 1, var, bg=row_bg, align='left',
              bold=(worst in IS_FAIL), sz=10)
        _cell(ws, data_row, 2, grp, bg=row_bg, sz=9, fg='555555')

        for ci, cid in enumerate(cust_ids):
            res_map = {r['col']: r for r in all_results.get(cid, [])}
            r = res_map.get(var, {'status': 'both_nan', 'ref': None, 'code': None, 'pct_diff': None})
            base = base_cols + ci * per_cust + 1

            bg_key, fg_key, _ = STATUS_STYLE.get(r['status'], ('white', 'dark', ''))
            bg = PAL[bg_key]; fg = PAL[fg_key]

            _cell(ws, data_row, base,   fmt_val(r['ref']),  bg=bg, fg=PAL['dark'], sz=10)
            _cell(ws, data_row, base+1, fmt_val(r['code']), bg=bg, fg=PAL['dark'], sz=10)
            _status_cell(ws, data_row, base+2, r['status'])
            pd_str = f"{r['pct_diff']:.2f}%" if r['pct_diff'] is not None else '—'
            _cell(ws, data_row, base+3, pd_str, bg=bg, fg=fg, sz=9)

        data_row += 1


# ── Sheet 3: Mismatches ───────────────────────────────────────────────────────

def build_mismatches(wb, all_results: dict[str, list], cust_ids: list[str]):
    ws = wb.create_sheet('✘ Mismatches')
    ws.sheet_view.showGridLines = False
    _col_w(ws, [40, 14, 16, 16, 12, 12, 45])

    _header(ws, 1, 1, 'Variables Requiring Attention — Mismatches, Warnings & Known Issues',
            span=7, sz=12)

    hdrs = ['Variable', 'Customer', 'Reference', 'Code Output', 'Status', '% Diff', 'Note / Root Cause']
    ws.row_dimensions[2].height = 20
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, 2, ci, h, bg=PAL['head_bg'], fg=PAL['head_fg'], bold=True, sz=10)

    data_row = 3
    non_match_statuses = {'mismatch', 'known_issue', 'ref_nan', 'code_nan'}

    for cid in cust_ids:
        for r in sorted(all_results[cid], key=lambda x: STATUS_RANK.get(x['status'], 0), reverse=True):
            if r['status'] not in non_match_statuses:
                continue
            bg_key, fg_key, lbl = STATUS_STYLE.get(r['status'], ('white', 'dark', r['status']))
            bg = PAL[bg_key]; fg = PAL[fg_key]
            ws.row_dimensions[data_row].height = 17

            _cell(ws, data_row, 1, r['col'],    bg=bg, align='left', bold=True, sz=10)
            _cell(ws, data_row, 2, f"cust {cid}", bg=bg, sz=10)
            _cell(ws, data_row, 3, fmt_val(r['ref']),  bg=bg, sz=10)
            _cell(ws, data_row, 4, fmt_val(r['code']), bg=bg, sz=10)
            _status_cell(ws, data_row, 5, r['status'])
            pd_str = f"{r['pct_diff']:.2f}%" if r['pct_diff'] is not None else '—'
            _cell(ws, data_row, 6, pd_str, bg=bg, fg=fg, sz=9)
            _cell(ws, data_row, 7, r.get('note',''), bg=PAL['white'], fg='555555',
                  align='left', sz=9, wrap=True)
            data_row += 1

    if data_row == 3:
        ws.merge_cells(f'A3:G3')
        _cell(ws, 3, 1, '✔ No mismatches or warnings found — all variables match!',
              bg=PAL['exact_bg'], fg=PAL['exact_fg'], bold=True, sz=12, span=7)


# ── Sheet 4: Column Audit ─────────────────────────────────────────────────────

def build_column_audit(wb, ref_cols: list[str], code_cols: list[str], label: str):
    ws = wb.create_sheet('🔍 Column Audit')
    ws.sheet_view.showGridLines = False
    _col_w(ws, [45, 20, 50])

    _header(ws, 1, 1, f'Column Audit — {label}', span=3, sz=12)

    audit = audit_columns(ref_cols, code_cols)

    sections = [
        ('Missing from Code Output (in Reference but not in pipeline CSV)',
         audit['missing'], PAL['fail_bg'], PAL['fail_fg'],
         'Add this column to the output_cols list in run_pipeline()'),
        ('Extra in Code Output (not in Reference)',
         audit['extra'], PAL['warn_bg'], PAL['warn_fg'],
         'Column computed but not expected — verify output_cols'),
        ('Duplicate Column Names in Code Output',
         audit['dupes_code'], PAL['known_bg'], PAL['known_fg'],
         'Column appears more than once — will cause value misalignment!'),
        ('Matched Columns (present in both)',
         audit['common'], PAL['exact_bg'], PAL['exact_fg'],
         ''),
    ]

    row = 2
    for title, items, bg, fg, hint in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, f'{title}  ({len(items)})')
        c.font = Font(bold=True, color=fg, name='Calibri', size=10)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = _thin()
        ws.row_dimensions[row].height = 20
        row += 1

        for item in items:
            ws.row_dimensions[row].height = 16
            _cell(ws, row, 1, item, bg=bg, align='left', sz=10)
            _cell(ws, row, 2, '',  bg=bg)
            _cell(ws, row, 3, hint, bg=PAL['white'], fg='555555', align='left', sz=9)
            row += 1

        if not items:
            _cell(ws, row, 1, '(none)', bg=PAL['light'], fg='888888', align='left', sz=9)
            row += 1

        row += 1  # spacer


# =============================================================================
# SECTION 6 — PYSPARK COMPARISON (for large files)
# =============================================================================

def spark_compare(ref_path: str, pipeline_path: str,
                  exact_tol: float, near_pct: float) -> dict[str, list]:
    """
    Use PySpark to compare large reference/pipeline files.
    Falls back to pandas for small files.
    Returns same structure as pandas compare.
    """
    if not SPARK_AVAILABLE:
        print("[INFO] PySpark not available — using pandas comparison")
        return None

    print("[INFO] Building Spark session for large-file comparison...")
    spark = (SparkSession.builder
             .master("local[*]")
             .appName("ScorecardValidator")
             .config("spark.driver.memory", "4g")
             .config("spark.sql.execution.arrow.pyspark.enabled", "false")
             .config("spark.sql.shuffle.partitions", "4")
             .getOrCreate())
    spark.sparkContext.setLogLevel("ERROR")

    def _load_spark(path):
        ext = os.path.splitext(path)[-1].lower()
        if ext in ('.xlsx', '.xls'):
            pdf = pd.read_excel(path)
            return spark.createDataFrame(pdf.astype(str))
        return spark.read.option("header", True).option("inferSchema", True).csv(path)

    ref_sdf  = _load_spark(ref_path)
    code_sdf = _load_spark(pipeline_path)

    ref_cols  = [c for c in ref_sdf.columns  if c != 'cust_id']
    code_cols = [c for c in code_sdf.columns if c != 'cust_id']
    common    = [c for c in ref_cols if c in set(code_cols)]

    # Rename code columns to avoid collision
    code_renamed = code_sdf.select(
        ['cust_id'] + [F.col(c).alias(f'_code_{c}') for c in common]
    )
    joined = ref_sdf.join(code_renamed, on='cust_id', how='outer')

    rows = joined.collect()
    all_results = {}
    for row in rows:
        cid = str(row['cust_id'])
        results = []
        for col in common:
            rv = row[col]
            cv = row[f'_code_{col}']
            try: rv = float(rv) if rv is not None else None
            except: pass
            try: cv = float(cv) if cv is not None else None
            except: pass
            r = classify_pair(rv, cv, col, exact_tol, near_pct)
            r['col'] = col
            r['group'] = next((g for g, cols in VARIABLE_GROUPS.items() if col in cols), 'Other')
            results.append(r)
        all_results[cid] = results

    spark.stop()
    return all_results


# =============================================================================
# SECTION 7 — MAIN ORCHESTRATOR
# =============================================================================

def run_validation(reference_path: str,
                   pipeline_path: str,
                   output_path: str,
                   label: str = "Pipeline Validation",
                   exact_tol: float = EXACT_TOL_DEFAULT,
                   near_pct: float = NEAR_PCT_DEFAULT,
                   use_spark: bool = False,
                   skip_cols: set | None = None) -> dict:
    """
    Full validation pipeline. Returns summary dict.

    skip_cols : optional set of extra column names to exclude from comparison
                on top of the built-in PIPELINE_ONLY_COLS
                (scorecard_name, final_score, score_breakdown,
                 is_eligible_for_st2, is_eligible_for_st3).
    """
    print(f"\n{'='*65}")
    print(f"  CONSUMER SCORECARD VALIDATOR")
    print(f"{'='*65}")
    print(f"  Reference :  {reference_path}")
    print(f"  Pipeline  :  {pipeline_path}")
    print(f"  Output    :  {output_path}")
    print(f"  Label     :  {label}")
    print(f"  Exact tol :  {exact_tol}")
    print(f"  Near %    :  {near_pct}%")
    print(f"{'='*65}\n")

    run_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Resolve full skip set: built-in pipeline cols + any user-supplied extras
    full_skip = PIPELINE_ONLY_COLS | (skip_cols or set())

    print(f"\n{'='*65}")
    print(f"  CONSUMER SCORECARD VALIDATOR")
    print(f"{'='*65}")
    print(f"  Reference   :  {reference_path}")
    print(f"  Pipeline    :  {pipeline_path}")
    print(f"  Output      :  {output_path}")
    print(f"  Label       :  {label}")
    print(f"  Exact tol   :  {exact_tol}")
    print(f"  Near %      :  {near_pct}%")
    print(f"  Always skip :  {sorted(PIPELINE_ONLY_COLS)}")
    if skip_cols:
        print(f"  Also skip   :  {sorted(skip_cols)}")
    print(f"{'='*65}\n")

    # ── Load data ─────────────────────────────────────────────────────────────
    print("Loading reference data...")
    ref_data  = load_reference(reference_path)
    print(f"  → {len(ref_data)} customers: {sorted(ref_data.keys())}")

    print("Loading pipeline output...")
    code_data = load_pipeline_output(pipeline_path, skip_cols=full_skip)
    print(f"  → {len(code_data)} customers: {sorted(code_data.keys())}")

    # ── Column audit ──────────────────────────────────────────────────────────
    # Exclude skipped cols from audit so they don't show as "missing from code"
    ref_all_cols  = sorted(c for s in ref_data.values()  for c in s.index)
    code_all_cols = sorted(c for s in code_data.values() for c in s.index
                           if c not in full_skip)
    audit = audit_columns(ref_all_cols, code_all_cols)
    if audit['missing']:
        print(f"\n  [WARN] {len(audit['missing'])} columns missing from pipeline output:")
        for c in audit['missing'][:10]: print(f"    - {c}")
    if audit['dupes_code']:
        print(f"\n  [WARN] Duplicate columns in pipeline output: {audit['dupes_code']}")

    # ── Compare ───────────────────────────────────────────────────────────────
    all_results: dict[str, list] = {}

    if use_spark and SPARK_AVAILABLE:
        print("\nRunning Spark comparison...")
        all_results = spark_compare(reference_path, pipeline_path, exact_tol, near_pct) or {}
    
    if not all_results:
        print("\nRunning pandas comparison...")
        common_custs = sorted(set(ref_data.keys()) & set(code_data.keys()))
        ref_only     = sorted(set(ref_data.keys()) - set(code_data.keys()))
        code_only    = sorted(set(code_data.keys()) - set(ref_data.keys()))
        if ref_only:
            print(f"  [WARN] Customers in reference only (no pipeline output): {ref_only}")
        if code_only:
            print(f"  [WARN] Customers in pipeline only (no reference): {code_only}")

        for cid in common_custs:
            print(f"  Comparing cust {cid}...")
            results = compare_customer(ref_data[cid], code_data[cid], exact_tol, near_pct)
            all_results[cid] = results

    # ── Print console summary ─────────────────────────────────────────────────
    print(f"\n{'─'*55}")
    print(f"  {'Customer':<20} {'Match':>8} {'Total':>7} {'Pct':>8}  {'Fail':>6}  {'Warn':>6}")
    print(f"{'─'*55}")
    for cid in sorted(all_results.keys()):
        s = summarise(all_results[cid])
        bar = '█' * int(s['pct'] / 5) + '░' * (20 - int(s['pct'] / 5))
        print(f"  cust {cid:<15} {s['match']:>8} {s['total']:>7} {s['pct']:>7.1f}%  "
              f"{s['mismatch']:>6}  {s['ref_nan']+s['code_nan']:>6}  [{bar}]")
    print(f"{'─'*55}")

    # ── Build Excel report ────────────────────────────────────────────────────
    if not OPENPYXL_AVAILABLE:
        print("\n[ERROR] openpyxl not installed — cannot write Excel report.")
        print("        Run: pip install openpyxl")
        return all_results

    print(f"\nBuilding Excel report → {output_path}")
    wb = Workbook()
    cust_ids = sorted(all_results.keys())

    build_dashboard(wb, all_results, label, run_ts)
    build_detail(wb, all_results, cust_ids)
    build_mismatches(wb, all_results, cust_ids)
    build_column_audit(wb, ref_all_cols, code_all_cols, label)

    wb.save(output_path)
    print(f"  ✅ Saved: {output_path}")

    return all_results


# =============================================================================
# SECTION 8 — CLI ENTRY POINT
# =============================================================================

def _parse_args():
    p = argparse.ArgumentParser(
        description="Consumer Scorecard V4 — PySpark Validation Script",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
See the header of this script for full usage examples and documentation.
The 5 pipeline scoring columns (scorecard_name, final_score, score_breakdown,
is_eligible_for_st2, is_eligible_for_st3) are always skipped automatically.
Use --skip-cols to exclude additional columns beyond those 5.
""")
    p.add_argument('--reference',  required=True,  help='Reference ground-truth CSV or Excel')
    p.add_argument('--pipeline',   required=True,  help='Pipeline output CSV or Excel')
    p.add_argument('--output',     default='validation_report.xlsx',
                                                   help='Output Excel report path')
    p.add_argument('--label',      default='Scorecard Validation',
                                                   help='Label shown in the report')
    p.add_argument('--exact-tol',  type=float, default=EXACT_TOL_DEFAULT,
                                                   help=f'Exact match tolerance (default {EXACT_TOL_DEFAULT})')
    p.add_argument('--near-tol',   type=float, default=NEAR_PCT_DEFAULT,
                                                   help=f'Near match %% tolerance (default {NEAR_PCT_DEFAULT})')
    p.add_argument('--spark',      action='store_true',
                                                   help='Use PySpark for large-file comparison')
    p.add_argument('--skip-cols',  nargs='+', default=[],
                                                   help=(
                                                       'Extra column names to exclude from comparison '
                                                       '(space-separated). The 5 pipeline scoring columns '
                                                       '(scorecard_name, final_score, score_breakdown, '
                                                       'is_eligible_for_st2, is_eligible_for_st3) are '
                                                       'always skipped automatically.'
                                                   ))
    return p.parse_args()


if __name__ == '__main__':
    args = _parse_args()
    # import pdb;pdb.set_trace()
    for path, name in [(args.reference, 'reference'), (args.pipeline, 'pipeline')]:
        if not os.path.exists(path):
            print(f"[ERROR] {name} file not found: {path}")
            sys.exit(1)

    run_validation(
        reference_path = args.reference,
        pipeline_path  = args.pipeline,
        output_path    = args.output,
        label          = args.label,
        exact_tol      = args.exact_tol,
        near_pct       = args.near_tol,
        use_spark      = args.spark,
        skip_cols      = set(args.skip_cols) if args.skip_cols else None,
    )

    # Step 1: ÷
# Step 2: run with your two files

# python validation.py --reference ground_truth.csv --pipeline scorecard.csv --output validation_report.xlsx --label "Scorecard Validation"